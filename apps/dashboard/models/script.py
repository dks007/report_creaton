import openpyxl
import os
from django.db import transaction
from django.contrib.auth.models import User
from apps.dashboard.models.masters import (
    CustomerMaster,
    IndustryMaster,
    CSMMaster,
    PSMMaster,
    SDMMaster,
    CustomerMapping,
    LogoMaster,
    StatusMaster
)

def get_or_create_logo(logo_url):
    if logo_url:
        logo_file_name = os.path.basename(logo_url)
        logo, _ = LogoMaster.objects.get_or_create(
            logo_url=logo_url,
            defaults={'logo_file_name': logo_file_name}
        )
        return logo
    return None

@transaction.atomic
def insert_from_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb[wb.sheetnames[0]]  # Use the first sheet

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Unpack the first 8 values from the row. The last two columns are ignored with `*rest`
            customer_id, customer_name, industry_name, sub_industry, csm_name, psm_name, sdm_name, logo_url, *rest = row

            industry, _ = IndustryMaster.objects.get_or_create(
                industry_type_name=industry_name) if industry_name else (None, False)

            # Now handle Sub Industry here if you need to associate it with the IndustryMaster or another model
            # ...

            csm, _ = CSMMaster.objects.get_or_create(csm_name=csm_name) if csm_name else (None, False)
            psm, _ = PSMMaster.objects.get_or_create(psm_name=psm_name) if psm_name else (None, False)
            sdm, _ = SDMMaster.objects.get_or_create(sdm_name=sdm_name) if sdm_name else (None, False)
            logo = get_or_create_logo(logo_url)

            customer, _ = CustomerMaster.objects.get_or_create(
                customer_id=customer_id,
                defaults={
                    'customer_name': customer_name,
                    'created_by': User.objects.first(),  # Adjust this as necessary
                    'updated_by': User.objects.first(),  # Adjust this as necessary
                    'status': StatusMaster.objects.first()  # Adjust this as necessary
                }
            )

            # Create CustomerMapping with all the collected master data
            CustomerMapping.objects.create(
                customer=customer,
                industry=industry,
                csm=csm,
                psm=psm,
                sdm=sdm,
                logo=logo,
                # Additional fields from your models if needed
            )

             

        return True

    except FileNotFoundError:
        print(f"File not found: {filename}")
        return False
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"Invalid Excel file: {filename}")
        return False
    except Exception as e:
        print(f"An error occurred: {e}")
        return False

# Specify the path to your Excel file
filename = 'E:\\IFS_BACKEND\\success_tool_backend_local\\report_creaton\\apps\\dashboard\\models\\master_data_13march.xlsx'
if insert_from_excel(filename):
    print("Data insertion successful.")
else:
    print("Data insertion failed.")
