from openpyxl import load_workbook
import pymysql

# Replace these with your MySQL database details
db_host = '************'
db_user = '********'
db_password = '*********'
db_name = '********'

# Replace this with the path to your Excel file
excel_file_path = '****************************'

# Specify the sheet name you want to read
sheet_name = '************'

# Create a MySQL connection
connection = pymysql.connect(
    host=db_host,
    user=db_user,
    password=db_password,
    database=db_name,
    port=5432
)

# Create a cursor object
cursor = connection.cursor()

# Read data from the Excel file
workbook = load_workbook(excel_file_path, read_only=True)
sheet = workbook[sheet_name]

# Extract and insert data into MySQL table
table_name = '*****************'
for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming data starts from row 2
    # Assuming your table structure is (id, sdo_name, sdo_email, created_by, updated_by, created_date, updated_date, status)
    sdo_name = row[1] #row index need to insert
    sdo_email = row[2] if len(row) > 2 and row[2] is not None else 'NA'
    # created_by =1
    # updated_by =1
    # status='1'

    # Check if sdo_name is None or empty
    if sdo_name is None or sdo_name == '':
        print(f"Skipping row due to missing or empty 'sdo_name'.")
        continue

    # Construct the SQL INSERT statement with placeholders
    sql = f"INSERT INTO {table_name} (sdo_name, sdo_email) VALUES (%s, %s)"

    # Execute the SQL statement with parameterized values
    cursor.execute(sql, (sdo_name, sdo_email))

# Commit the changes and close the connections
connection.commit()
cursor.close()
connection.close()

print(f"Data from sheet '{sheet_name}' successfully loaded into MySQL table '{table_name}'.")

